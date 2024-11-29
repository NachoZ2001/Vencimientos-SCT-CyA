from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import win32com.client as win32
from openpyxl.utils import get_column_letter
import pandas as pd
import time
import pyautogui
import os
import glob
import random
import xlwings as xw
import pdfkit
import os
import pandas as pd
import glob

# Obtener la ruta base del directorio donde está el script
base_dir = os.path.dirname(os.path.abspath(__file__))

# Definir rutas a las carpetas y archivos
input_folder_excel = os.path.join(base_dir, "data", "input", "Deudas")
output_folder_csv = os.path.join(base_dir, "data", "input", "DeudasCSV")
output_file_csv = os.path.join(base_dir, "data", "Resumen_deudas.csv")
output_file_xlsx = os.path.join(base_dir, "data", "Resumen_deudas.xlsx")

# Leer el archivo Excel
input_excel_clientes = os.path.join(base_dir, "data", "input", "clientes.xlsx")
df = pd.read_excel(input_excel_clientes)

# Suposición de nombres de columnas
cuit_login_list = df['CUIT para ingresar'].tolist()
cuit_represent_list = df['CUIT representado'].tolist()
password_list = df['Contraseña'].tolist()
download_list = df['Ubicacion descarga'].tolist()
posterior_list = df['Posterior'].tolist()
anterior_list = df['Anterior'].tolist()
clientes_list = df['Cliente'].tolist()

output_folder_pdf = os.path.join(base_dir, "data", "Reportes")
imagen = os.path.join(base_dir, "data", "imagen.png")

def forzar_guardado_excel(excel_file):
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(excel_file)
        wb.Save()
        wb.Close(False)
    except Exception as e:
        print(f"Error forzando guardado en {excel_file}: {e}")
    finally:
        excel.Quit()

def ajustar_diseno_excel(ws):
    """
    Ajusta el diseño del archivo Excel para que todo el contenido (imagen y tabla) 
    quepa en una sola página PDF.
    """
    # Configurar ajuste de página para que quepa todo en una página
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = "landscape"  # Apaisado
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

def procesar_excel(excel_file, output_pdf, imagen):
    try:
        # Cargar el archivo Excel con pandas
        df = pd.read_excel(excel_file)

        # Filtrar por "Periodo fiscal" y "Impuesto"
        df_filtrado = df[
            (df['Impuesto'].str.contains('ganancias sociedades', case=False, na=False))
        ]

        # Verificar si la tabla está vacía
        if df_filtrado.shape[0] == 0:
            output_pdf = output_pdf.replace(".pdf", " - vacio.pdf")

        # Eliminar las columnas innecesarias
        columnas_a_eliminar = ['Concepto / Subconcepto', 'Int. resarcitorios', 'Int. punitorios']
        for columna in columnas_a_eliminar:
            if columna in df.columns:
                df_filtrado = df_filtrado.drop(columna, axis=1)

        # Guardar el DataFrame filtrado en el archivo Excel
        df_filtrado.to_excel(excel_file, index=False)

        # Cargar el archivo para aplicar formato con openpyxl
        wb = load_workbook(excel_file)
        ws = wb.active

        # Insertar filas adicionales para una nueva imagen
        ws.insert_rows(1, amount=7)

        # Agregar una imagen encima del encabezado (A1)
        # Obtener el ancho combinado de la tabla
        ultima_columna = ws.max_column
        ultima_letra_columna = get_column_letter(ultima_columna)

        # Insertar la imagen
        img = ExcelImage(imagen)
        # Ajustar el tamaño de la imagen (puedes personalizar estos valores)
        img.width = ws.column_dimensions['A'].width * ultima_columna * 6  # Ajustar al ancho combinado
        img.height = 120  # Altura fija
        # Agregar la imagen a la hoja (posicionándola dentro del rango combinado)
        ws.add_image(img, 'A1')

        # Insertar filas adicionales para una nueva imagen
        ws.insert_rows(7, amount=1)

        # Fila donde se agregará el texto
        fila_texto = 8

        # Obtener el número de columnas ocupadas por la tabla
        ultima_columna = ws.max_column
        ultima_letra_columna = get_column_letter(ultima_columna)

        # Combinar celdas en la fila de separación
        ws.merge_cells(f'A{fila_texto}:{ultima_letra_columna}{fila_texto}')

        # Establecer el texto en la celda combinada
        celda_texto = ws[f'A{fila_texto}']
        celda_texto.value = "Cronograma de anticipos"

        # Aplicar formato centrado y en negrita
        celda_texto.alignment = Alignment(horizontal='center', vertical='center')
        celda_texto.font = Font(bold=True, size=20)

        # Cambiar el color del encabezado a lila
        header_fill = PatternFill(start_color="AA0EAA", end_color="AA0EAA", fill_type="solid")
        for cell in ws[9]:
            cell.fill = header_fill

        # Ajustar el ancho de las columnas automáticamente, pero individualmente
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            # Ajuste del ancho de cada columna basado en el contenido más largo
            adjusted_width = (max_length + 2) * 1.2  # 1.2 para un poco de margen adicional
            ws.column_dimensions[column].width = adjusted_width

        # Centrar el contenido de todas las celdas
        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guardar los cambios
        wb.save(excel_file)

        # Convertir el archivo Excel a PDF con pywin32
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(excel_file)

        print("Configurando área de impresión...")
        ws = wb.Worksheets(1)

        # Definir el rango del área de impresión manualmente
        last_row = ws.UsedRange.Rows.Count
        last_col = ws.UsedRange.Columns.Count
        ws.PageSetup.PrintArea = f"A1:{get_column_letter(last_col)}{last_row + 8}"  # Incluir imagen y tabla

        # Ajustar a una página
        ws.PageSetup.Orientation = 2  # Paisaje
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = 1

        # Configurar centrado en la página
        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = False  # Verticalmente opcional, según el diseño

        # Configurar márgenes
        ws.PageSetup.LeftMargin = 0.25
        ws.PageSetup.RightMargin = 0.25
        ws.PageSetup.TopMargin = 0.5
        ws.PageSetup.BottomMargin = 0.5

        print("Guardando como PDF...")
        wb.ExportAsFixedFormat(0, output_pdf)  # 0 indica formato PDF
        wb.Close(False)
        print(f"Archivo convertido a PDF: {output_pdf}")

    except Exception as e:
        print(f"Error al procesar {excel_file}: {e}")
    finally:
        if 'excel' in locals():
            excel.Quit()

# Recorrer todos los archivos Excel en la carpeta
for excel_file in glob.glob(os.path.join(input_folder_excel, "*.xlsx")):
    try:
        # Forzar guardado para evitar problemas con archivos corruptos o no calculados
        forzar_guardado_excel(excel_file)

        # Obtener el nombre base del archivo para usarlo en el nombre del PDF
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_pdf = os.path.join(output_folder_pdf, f"{base_name}.pdf")
        
        # Llamar a la función para procesar el archivo Excel y generar el PDF
        procesar_excel(excel_file, output_pdf, imagen)
        
        print(f"Archivo {excel_file} procesado y guardado como {output_pdf}")
    
    except Exception as e:
        print(f"Error al procesar {excel_file}: {e}")
